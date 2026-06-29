import os
import uuid
import aiofiles
from pathlib import Path
from typing import Optional, Tuple
from fastapi import UploadFile, HTTPException, status

from app.core.config import settings


async def save_upload_file(file: UploadFile) -> Tuple[str, str, int]:
    """Save uploaded file and return (filename, filepath, size)."""
    upload_dir = Path(settings.UPLOAD_DIR)
    upload_dir.mkdir(parents=True, exist_ok=True)

    ext = Path(file.filename).suffix.lower().lstrip(".")
    if ext not in settings.ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File type '{ext}' not allowed. Allowed: {', '.join(settings.ALLOWED_EXTENSIONS)}"
        )

    unique_filename = f"{uuid.uuid4().hex}_{file.filename}"
    file_path = upload_dir / unique_filename

    content = await file.read()
    file_size = len(content)

    if file_size > settings.MAX_FILE_SIZE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File too large. Maximum size: {settings.MAX_FILE_SIZE_MB}MB"
        )

    async with aiofiles.open(file_path, "wb") as f:
        await f.write(content)

    return unique_filename, str(file_path), file_size


async def extract_text_from_file(file_path: str, file_type: str) -> Optional[str]:
    """Extract text content from various file types."""
    try:
        if file_type == "txt":
            return await _extract_txt(file_path)
        elif file_type == "pdf":
            return await _extract_pdf(file_path)
        elif file_type == "docx":
            return await _extract_docx(file_path)
        elif file_type in ("xlsx", "xls"):
            return await _extract_excel(file_path)
        elif file_type == "csv":
            return await _extract_csv(file_path)
        return None
    except Exception as e:
        print(f"Error extracting text from {file_path}: {e}")
        return None


async def _extract_txt(file_path: str) -> str:
    async with aiofiles.open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return await f.read()


async def _extract_pdf(file_path: str) -> str:
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
        return "\n\n".join(text_parts)
    except ImportError:
        try:
            import PyPDF2
            text_parts = []
            with open(file_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
            return "\n\n".join(text_parts)
        except Exception as e:
            return f"[PDF text extraction failed: {e}]"


async def _extract_docx(file_path: str) -> str:
    try:
        from docx import Document
        doc = Document(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n\n".join(paragraphs)
    except Exception as e:
        return f"[DOCX extraction failed: {e}]"


async def _extract_excel(file_path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}")
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(cell.strip() for cell in row_data):
                    rows.append(" | ".join(row_data))
            text_parts.extend(rows[:100])  # Max 100 rows per sheet
        return "\n".join(text_parts)
    except Exception as e:
        return f"[Excel extraction failed: {e}]"


async def _extract_csv(file_path: str) -> str:
    try:
        import csv
        rows = []
        async with aiofiles.open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            content = await f.read()
        reader = csv.reader(content.splitlines())
        for i, row in enumerate(reader):
            if i > 200:  # Max 200 rows
                rows.append("... (truncated)")
                break
            rows.append(" | ".join(row))
        return "\n".join(rows)
    except Exception as e:
        return f"[CSV extraction failed: {e}]"


def delete_file(file_path: str) -> bool:
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            return True
        return False
    except Exception:
        return False
